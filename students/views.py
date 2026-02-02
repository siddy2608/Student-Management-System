from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.db.models import Q, Count, Avg, Sum
from django.db.models.functions import TruncMonth
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import timedelta
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import Student, Course, Department, Enrollment, Attendance, Fee, Announcement
from .forms import (
    StudentForm, CourseForm, DepartmentForm, EnrollmentForm, 
    AttendanceForm, BulkAttendanceForm, FeeForm, AnnouncementForm, StudentFilterForm,
    SignUpForm
)


# ============== Authentication Views ==============

def login_view(request):
    """User login view."""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Welcome back, {username}!')
                next_url = request.GET.get('next', 'dashboard')
                return redirect(next_url)
    else:
        form = AuthenticationForm()
    
    return render(request, 'students/auth/login.html', {'form': form})


def signup_view(request):
    """User registration view."""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f'Welcome to StudentHub, {user.username}! Your account has been created.')
            return redirect('dashboard')
    else:
        form = SignUpForm()
    
    return render(request, 'students/auth/signup.html', {'form': form})


def logout_view(request):
    """User logout view."""
    logout(request)
    messages.info(request, 'You have been logged out.')
    return redirect('login')


# ============== Dashboard Views ==============

@login_required
def dashboard(request):
    """Main dashboard with analytics."""
    # Basic counts
    total_students = Student.objects.count()
    active_students = Student.objects.filter(is_active=True).count()
    total_courses = Course.objects.filter(is_active=True).count()
    total_departments = Department.objects.count()
    
    # Recent students
    recent_students = Student.objects.order_by('-created_at')[:5]
    
    # Department-wise student distribution
    dept_distribution = Department.objects.annotate(
        student_count=Count('students')
    ).values('name', 'student_count')
    
    # Monthly enrollment trend (last 6 months)
    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_enrollments = Student.objects.filter(
        admission_date__gte=six_months_ago
    ).annotate(
        month=TruncMonth('admission_date')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    # GPA distribution
    gpa_ranges = {
        '3.5 - 4.0': Student.objects.filter(gpa__gte=3.5).count(),
        '3.0 - 3.49': Student.objects.filter(gpa__gte=3.0, gpa__lt=3.5).count(),
        '2.5 - 2.99': Student.objects.filter(gpa__gte=2.5, gpa__lt=3.0).count(),
        '2.0 - 2.49': Student.objects.filter(gpa__gte=2.0, gpa__lt=2.5).count(),
        'Below 2.0': Student.objects.filter(gpa__lt=2.0).count(),
    }
    
    # Attendance overview (today)
    today = timezone.now().date()
    today_attendance = Attendance.objects.filter(date=today)
    present_today = today_attendance.filter(status='P').count()
    absent_today = today_attendance.filter(status='A').count()
    
    # Pending fees
    pending_fees = Fee.objects.filter(status='PEN').aggregate(
        total=Sum('amount'),
        count=Count('id')
    )
    
    # Active announcements
    announcements = Announcement.objects.filter(
        is_active=True
    ).order_by('-priority', '-created_at')[:3]
    
    context = {
        'total_students': total_students,
        'active_students': active_students,
        'total_courses': total_courses,
        'total_departments': total_departments,
        'recent_students': recent_students,
        'dept_distribution': list(dept_distribution),
        'monthly_enrollments': list(monthly_enrollments),
        'gpa_ranges': gpa_ranges,
        'present_today': present_today,
        'absent_today': absent_today,
        'pending_fees': pending_fees,
        'announcements': announcements,
    }
    return render(request, 'students/dashboard.html', context)


# ============== Student Views ==============

@login_required
def student_list(request):
    """Display list of students with filtering, sorting, and pagination."""
    students = Student.objects.select_related('department').all()
    
    # Search
    query = request.GET.get('q', '')
    if query:
        students = students.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query) |
            Q(student_id__icontains=query)
        )
    
    # Filter by department
    department_id = request.GET.get('department', '')
    if department_id:
        students = students.filter(department_id=department_id)
    
    # Filter by status
    status = request.GET.get('status', '')
    if status == 'active':
        students = students.filter(is_active=True)
    elif status == 'inactive':
        students = students.filter(is_active=False)
    
    # Filter by semester
    semester = request.GET.get('semester', '')
    if semester:
        students = students.filter(current_semester=semester)
    
    # Sorting
    sort_by = request.GET.get('sort', '-created_at')
    valid_sorts = ['student_id', '-student_id', 'first_name', '-first_name', 
                   'gpa', '-gpa', 'created_at', '-created_at']
    if sort_by in valid_sorts:
        students = students.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(students, 10)
    page = request.GET.get('page', 1)
    students_page = paginator.get_page(page)
    
    # Get all departments for filter dropdown
    departments = Department.objects.all()
    
    context = {
        'students': students_page,
        'query': query,
        'departments': departments,
        'current_department': department_id,
        'current_status': status,
        'current_semester': semester,
        'current_sort': sort_by,
        'total_count': paginator.count,
    }
    return render(request, 'students/student_list.html', context)


@login_required
def student_detail(request, pk):
    """Display detailed information about a student."""
    student = get_object_or_404(Student.objects.select_related('department'), pk=pk)
    
    # Get enrollments with courses
    enrollments = Enrollment.objects.filter(student=student).select_related('course')
    
    # Get attendance summary
    attendance_records = Attendance.objects.filter(student=student)
    attendance_summary = {
        'total': attendance_records.count(),
        'present': attendance_records.filter(status='P').count(),
        'absent': attendance_records.filter(status='A').count(),
        'late': attendance_records.filter(status='L').count(),
        'excused': attendance_records.filter(status='E').count(),
    }
    
    # Get fees
    fees = Fee.objects.filter(student=student).order_by('-due_date')[:5]
    fee_summary = Fee.objects.filter(student=student).aggregate(
        total_due=Sum('amount', filter=Q(status='PEN')),
        total_paid=Sum('amount', filter=Q(status='PAI'))
    )
    
    context = {
        'student': student,
        'enrollments': enrollments,
        'attendance_summary': attendance_summary,
        'fees': fees,
        'fee_summary': fee_summary,
    }
    return render(request, 'students/student_detail.html', context)


@login_required
def student_create(request):
    """Create a new student."""
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            student = form.save()
            messages.success(request, f'Student "{student.full_name}" created successfully!')
            return redirect('student_detail', pk=student.pk)
    else:
        form = StudentForm()
    
    return render(request, 'students/student_form.html', {
        'form': form,
        'title': 'Add New Student',
        'button_text': 'Create Student'
    })


@login_required
def student_update(request, pk):
    """Update an existing student."""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            student = form.save()
            messages.success(request, f'Student "{student.full_name}" updated successfully!')
            return redirect('student_detail', pk=student.pk)
    else:
        form = StudentForm(instance=student)
    
    return render(request, 'students/student_form.html', {
        'form': form,
        'student': student,
        'title': f'Edit: {student.full_name}',
        'button_text': 'Update Student'
    })


@login_required
def student_delete(request, pk):
    """Delete a student."""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        name = student.full_name
        student.delete()
        messages.success(request, f'Student "{name}" deleted successfully!')
        return redirect('student_list')
    
    return render(request, 'students/student_confirm_delete.html', {'student': student})


# ============== Course Views ==============

@login_required
def course_list(request):
    """Display list of courses."""
    courses = Course.objects.select_related('department').annotate(
        enrolled_students=Count('enrollments', filter=Q(enrollments__is_active=True))
    )
    
    # Search
    query = request.GET.get('q', '')
    if query:
        courses = courses.filter(
            Q(name__icontains=query) |
            Q(code__icontains=query) |
            Q(instructor__icontains=query)
        )
    
    # Filter by department
    department_id = request.GET.get('department', '')
    if department_id:
        courses = courses.filter(department_id=department_id)
    
    # Pagination
    paginator = Paginator(courses, 10)
    page = request.GET.get('page', 1)
    courses_page = paginator.get_page(page)
    
    departments = Department.objects.all()
    
    context = {
        'courses': courses_page,
        'query': query,
        'departments': departments,
        'current_department': department_id,
    }
    return render(request, 'students/course_list.html', context)


@login_required
def course_detail(request, pk):
    """Display course details with enrolled students."""
    course = get_object_or_404(Course.objects.select_related('department'), pk=pk)
    
    enrollments = Enrollment.objects.filter(course=course, is_active=True).select_related('student')
    
    # Grade distribution
    grade_dist = Enrollment.objects.filter(course=course, grade__isnull=False).values('grade').annotate(
        count=Count('id')
    )
    
    context = {
        'course': course,
        'enrollments': enrollments,
        'grade_distribution': list(grade_dist),
    }
    return render(request, 'students/course_detail.html', context)


@login_required
def course_create(request):
    """Create a new course."""
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save()
            messages.success(request, f'Course "{course.code}" created successfully!')
            return redirect('course_detail', pk=course.pk)
    else:
        form = CourseForm()
    
    return render(request, 'students/course_form.html', {
        'form': form,
        'title': 'Add New Course',
        'button_text': 'Create Course'
    })


@login_required
def course_update(request, pk):
    """Update an existing course."""
    course = get_object_or_404(Course, pk=pk)
    
    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            course = form.save()
            messages.success(request, f'Course "{course.code}" updated successfully!')
            return redirect('course_detail', pk=course.pk)
    else:
        form = CourseForm(instance=course)
    
    return render(request, 'students/course_form.html', {
        'form': form,
        'course': course,
        'title': f'Edit: {course.code}',
        'button_text': 'Update Course'
    })


@login_required
def course_delete(request, pk):
    """Delete a course."""
    course = get_object_or_404(Course, pk=pk)
    
    if request.method == 'POST':
        code = course.code
        course.delete()
        messages.success(request, f'Course "{code}" deleted successfully!')
        return redirect('course_list')
    
    return render(request, 'students/course_confirm_delete.html', {'course': course})


# ============== Department Views ==============

@login_required
def department_list(request):
    """Display list of departments."""
    departments = Department.objects.annotate(
        student_count=Count('students'),
        course_count=Count('courses')
    )
    
    context = {'departments': departments}
    return render(request, 'students/department_list.html', context)


@login_required
def department_create(request):
    """Create a new department."""
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            department = form.save()
            messages.success(request, f'Department "{department.name}" created successfully!')
            return redirect('department_list')
    else:
        form = DepartmentForm()
    
    return render(request, 'students/department_form.html', {
        'form': form,
        'title': 'Add New Department',
        'button_text': 'Create Department'
    })


@login_required
def department_update(request, pk):
    """Update an existing department."""
    department = get_object_or_404(Department, pk=pk)
    
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            department = form.save()
            messages.success(request, f'Department "{department.name}" updated successfully!')
            return redirect('department_list')
    else:
        form = DepartmentForm(instance=department)
    
    return render(request, 'students/department_form.html', {
        'form': form,
        'department': department,
        'title': f'Edit: {department.name}',
        'button_text': 'Update Department'
    })


@login_required  
def department_delete(request, pk):
    """Delete a department."""
    department = get_object_or_404(Department, pk=pk)
    
    if request.method == 'POST':
        name = department.name
        department.delete()
        messages.success(request, f'Department "{name}" deleted successfully!')
        return redirect('department_list')
    
    return render(request, 'students/department_confirm_delete.html', {'department': department})


# ============== Enrollment Views ==============

@login_required
def enrollment_create(request):
    """Enroll a student in a course."""
    if request.method == 'POST':
        form = EnrollmentForm(request.POST)
        if form.is_valid():
            enrollment = form.save()
            messages.success(request, f'Student enrolled in {enrollment.course.code} successfully!')
            return redirect('student_detail', pk=enrollment.student.pk)
    else:
        form = EnrollmentForm()
        # Pre-fill student if provided
        student_id = request.GET.get('student')
        if student_id:
            form.fields['student'].initial = student_id
    
    return render(request, 'students/enrollment_form.html', {
        'form': form,
        'title': 'Enroll Student in Course',
        'button_text': 'Enroll'
    })


@login_required
def enrollment_update(request, pk):
    """Update enrollment (add grade)."""
    enrollment = get_object_or_404(Enrollment, pk=pk)
    
    if request.method == 'POST':
        form = EnrollmentForm(request.POST, instance=enrollment)
        if form.is_valid():
            enrollment = form.save()
            messages.success(request, 'Enrollment updated successfully!')
            return redirect('student_detail', pk=enrollment.student.pk)
    else:
        form = EnrollmentForm(instance=enrollment)
    
    return render(request, 'students/enrollment_form.html', {
        'form': form,
        'title': f'Update Enrollment: {enrollment.student.full_name} - {enrollment.course.code}',
        'button_text': 'Update'
    })


# ============== Attendance Views ==============

@login_required
def attendance_list(request):
    """Display attendance records."""
    attendance_records = Attendance.objects.select_related('student', 'course').all()
    
    # Filter by date
    date = request.GET.get('date', '')
    if date:
        attendance_records = attendance_records.filter(date=date)
    else:
        # Default to today
        attendance_records = attendance_records.filter(date=timezone.now().date())
    
    # Filter by course
    course_id = request.GET.get('course', '')
    if course_id:
        attendance_records = attendance_records.filter(course_id=course_id)
    
    courses = Course.objects.filter(is_active=True)
    
    context = {
        'attendance_records': attendance_records,
        'courses': courses,
        'current_date': date or timezone.now().date().isoformat(),
        'current_course': course_id,
    }
    return render(request, 'students/attendance_list.html', context)


@login_required
def attendance_take(request):
    """Take attendance for a course."""
    if request.method == 'POST':
        course_id = request.POST.get('course')
        date = request.POST.get('date', timezone.now().date())
        course = get_object_or_404(Course, pk=course_id)
        
        # Get enrolled students
        enrollments = Enrollment.objects.filter(course=course, is_active=True)
        
        created_count = 0
        for enrollment in enrollments:
            status = request.POST.get(f'status_{enrollment.student.pk}', 'P')
            remarks = request.POST.get(f'remarks_{enrollment.student.pk}', '')
            
            attendance, created = Attendance.objects.update_or_create(
                student=enrollment.student,
                course=course,
                date=date,
                defaults={
                    'status': status,
                    'remarks': remarks,
                    'recorded_by': request.user
                }
            )
            if created:
                created_count += 1
        
        messages.success(request, f'Attendance recorded for {created_count} students!')
        return redirect('attendance_list')
    
    # GET request - show form
    course_id = request.GET.get('course', '')
    date = request.GET.get('date', timezone.now().date().isoformat())
    
    courses = Course.objects.filter(is_active=True)
    students = []
    
    if course_id:
        course = get_object_or_404(Course, pk=course_id)
        enrollments = Enrollment.objects.filter(course=course, is_active=True).select_related('student')
        
        # Get existing attendance for this date
        existing_attendance = {
            a.student_id: a for a in 
            Attendance.objects.filter(course=course, date=date)
        }
        
        for enrollment in enrollments:
            existing = existing_attendance.get(enrollment.student.pk)
            students.append({
                'student': enrollment.student,
                'status': existing.status if existing else 'P',
                'remarks': existing.remarks if existing else '',
            })
    
    context = {
        'courses': courses,
        'students': students,
        'current_course': course_id,
        'current_date': date,
    }
    return render(request, 'students/attendance_take.html', context)


# ============== Fee Views ==============

@login_required
def fee_list(request):
    """Display fee records."""
    fees = Fee.objects.select_related('student').all()
    
    # Filter by status
    status = request.GET.get('status', '')
    if status:
        fees = fees.filter(status=status)
    
    # Search by student
    query = request.GET.get('q', '')
    if query:
        fees = fees.filter(
            Q(student__first_name__icontains=query) |
            Q(student__last_name__icontains=query) |
            Q(student__student_id__icontains=query)
        )
    
    # Summary
    summary = Fee.objects.aggregate(
        total_pending=Sum('amount', filter=Q(status='PEN')),
        total_paid=Sum('amount', filter=Q(status='PAI')),
        total_overdue=Sum('amount', filter=Q(status='OVD'))
    )
    
    # Pagination
    paginator = Paginator(fees, 15)
    page = request.GET.get('page', 1)
    fees_page = paginator.get_page(page)
    
    context = {
        'fees': fees_page,
        'summary': summary,
        'query': query,
        'current_status': status,
    }
    return render(request, 'students/fee_list.html', context)


@login_required
def fee_create(request):
    """Create a new fee record."""
    if request.method == 'POST':
        form = FeeForm(request.POST)
        if form.is_valid():
            fee = form.save()
            messages.success(request, f'Fee record created for {fee.student.full_name}!')
            return redirect('fee_list')
    else:
        form = FeeForm()
        student_id = request.GET.get('student')
        if student_id:
            form.fields['student'].initial = student_id
    
    return render(request, 'students/fee_form.html', {
        'form': form,
        'title': 'Add Fee Record',
        'button_text': 'Create Fee'
    })


@login_required
def fee_update(request, pk):
    """Update a fee record."""
    fee = get_object_or_404(Fee, pk=pk)
    
    if request.method == 'POST':
        form = FeeForm(request.POST, instance=fee)
        if form.is_valid():
            fee = form.save()
            messages.success(request, 'Fee record updated successfully!')
            return redirect('fee_list')
    else:
        form = FeeForm(instance=fee)
    
    return render(request, 'students/fee_form.html', {
        'form': form,
        'fee': fee,
        'title': f'Update Fee: {fee.student.full_name}',
        'button_text': 'Update Fee'
    })


# ============== Export Views ==============

@login_required
def export_students_excel(request):
    """Export students to Excel."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Students'
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Student ID', 'First Name', 'Last Name', 'Email', 'Phone', 'Department', 
               'Semester', 'GPA', 'Status', 'Admission Date']
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    students = Student.objects.select_related('department').all()
    
    # Apply filters if any
    query = request.GET.get('q', '')
    if query:
        students = students.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )
    
    department_id = request.GET.get('department', '')
    if department_id:
        students = students.filter(department_id=department_id)
    
    for row, student in enumerate(students, 2):
        sheet.cell(row=row, column=1, value=student.student_id).border = thin_border
        sheet.cell(row=row, column=2, value=student.first_name).border = thin_border
        sheet.cell(row=row, column=3, value=student.last_name).border = thin_border
        sheet.cell(row=row, column=4, value=student.email).border = thin_border
        sheet.cell(row=row, column=5, value=student.phone or '').border = thin_border
        sheet.cell(row=row, column=6, value=student.department.name if student.department else '').border = thin_border
        sheet.cell(row=row, column=7, value=student.current_semester).border = thin_border
        sheet.cell(row=row, column=8, value=float(student.gpa)).border = thin_border
        sheet.cell(row=row, column=9, value='Active' if student.is_active else 'Inactive').border = thin_border
        sheet.cell(row=row, column=10, value=student.admission_date.strftime('%Y-%m-%d')).border = thin_border
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[chr(64 + col)].width = 15
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=students_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    workbook.save(response)
    
    return response


@login_required
def export_attendance_excel(request):
    """Export attendance to Excel."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Attendance'
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    
    headers = ['Date', 'Student ID', 'Student Name', 'Course', 'Status', 'Remarks']
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Filter
    date = request.GET.get('date', '')
    course_id = request.GET.get('course', '')
    
    attendance_records = Attendance.objects.select_related('student', 'course').all()
    
    if date:
        attendance_records = attendance_records.filter(date=date)
    if course_id:
        attendance_records = attendance_records.filter(course_id=course_id)
    
    for row, record in enumerate(attendance_records, 2):
        sheet.cell(row=row, column=1, value=record.date.strftime('%Y-%m-%d'))
        sheet.cell(row=row, column=2, value=record.student.student_id)
        sheet.cell(row=row, column=3, value=record.student.full_name)
        sheet.cell(row=row, column=4, value=record.course.code)
        sheet.cell(row=row, column=5, value=record.get_status_display())
        sheet.cell(row=row, column=6, value=record.remarks)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=attendance_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    workbook.save(response)
    
    return response


# ============== API Views ==============

@login_required
def api_dashboard_stats(request):
    """API endpoint for dashboard statistics."""
    data = {
        'total_students': Student.objects.count(),
        'active_students': Student.objects.filter(is_active=True).count(),
        'total_courses': Course.objects.filter(is_active=True).count(),
        'total_departments': Department.objects.count(),
    }
    return JsonResponse(data)


@login_required
def api_attendance_chart(request):
    """API endpoint for attendance chart data."""
    course_id = request.GET.get('course')
    days = int(request.GET.get('days', 7))
    
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=days)
    
    attendance = Attendance.objects.filter(date__range=[start_date, end_date])
    
    if course_id:
        attendance = attendance.filter(course_id=course_id)
    
    data = attendance.values('date').annotate(
        present=Count('id', filter=Q(status='P')),
        absent=Count('id', filter=Q(status='A')),
        late=Count('id', filter=Q(status='L'))
    ).order_by('date')
    
    return JsonResponse(list(data), safe=False)
